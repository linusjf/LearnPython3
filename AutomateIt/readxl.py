#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

workbook = openpyxl.load_workbook("myxlsx.xlsx")
print("Workbook Object:", workbook.sheetnames)
people = workbook["People"]
print("People sheet object:", people)
print("First cell Object:", people["A1"])
print("Other Cell Object:", people.cell(row=3, column=2))
print("First Name:", people["B2"].value, people["C2"].value)
